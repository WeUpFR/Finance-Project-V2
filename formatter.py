import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, Any

def format_and_export_excel(statements: Dict[str, pd.DataFrame], output_path: str, company_name: str = "Demerara Life (Trinidad & Tobago)"):
    """
    Applies Demerara-style formatting and exports statements to an Excel file.

    Args:
        statements (Dict[str, pd.DataFrame]): Dictionary of statement DataFrames.
        output_path (str): Path to save the output Excel file.
        company_name (str): The name of the company for the report header.
    """
    wb = Workbook()
    wb.remove(wb.active) # Remove default sheet

    # Define styles
    header_font = Font(name='Arial', size=14, bold=True)
    title_font = Font(name='Arial', size=12, bold=True)
    total_font = Font(name='Arial', size=11, bold=True)
    section_header_font = Font(name='Arial', size=11, bold=True)
    currency_format = 'TT$ #,##0.00'
    
    top_border = Border(top=Side(style='thin'))
    double_bottom_border = Border(bottom=Side(style='double'))
    
    for statement_name, df in statements.items():
        if df.empty or df.iloc[0,0].startswith('Cash flow statement requires'):
            continue
            
        ws = wb.create_sheet(title=statement_name[:31]) # Sheet title max 31 chars

        # --- Header ---
        ws.merge_cells('A1:B1')
        ws.cell(row=1, column=1, value=company_name).font = header_font
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:B2')
        ws.cell(row=2, column=1, value=statement_name).font = title_font
        ws.cell(row=2, column=1).alignment = Alignment(horizontal='center')

        ws.merge_cells('A3:B3')
        ws.cell(row=3, column=1, value="For the period ended March 31, 2025").font = title_font
        ws.cell(row=3, column=1).alignment = Alignment(horizontal='center')

        # --- Column Headers ---
        ws.cell(row=5, column=1, value="Description").font = total_font
        ws.cell(row=5, column=2, value="Amount").font = total_font
        ws.cell(row=5, column=2).alignment = Alignment(horizontal='right')

        # --- Data and Styling ---
        current_row = 6
        for _, row_data in df.iterrows():
            line_item = row_data.get('Line Item', '')
            amount = row_data.get('Amount')
            
            ws.cell(row=current_row, column=1, value=line_item)
            if pd.notna(amount):
                ws.cell(row=current_row, column=2, value=amount)
                ws.cell(row=current_row, column=2).number_format = currency_format

            ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='right', vertical='center')
            
            if row_data.get('is_header'):
                ws.cell(row=current_row, column=1).font = section_header_font
            if row_data.get('is_total'):
                ws.cell(row=current_row, column=1).font = total_font
                ws.cell(row=current_row, column=2).font = total_font
                ws.cell(row=current_row, column=2).border = top_border
            if row_data.get('is_final_total'):
                ws.cell(row=current_row, column=2).border = Border(top=Side(style='thin'), bottom=Side(style='double'))
            
            current_row += 1

        # Column widths
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 25
    
    try:
        wb.save(output_path)
        print(f"Successfully exported styled Excel report to {output_path}")
    except Exception as e:
        print(f"Error saving Excel file: {e}")
